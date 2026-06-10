#!/usr/bin/env python3
# ==========================================================================
# xlsx_report.py  -  Render the AI-produced findings JSON into the client
# vulnerability-report .xlsx (two sheets, matching the house template):
#
#   Sheet 1  "Vulnerabilities"        - one row per finding, severity-ranked
#   Sheet 2  "Attack Paths & Chains"  - realistic attacker chains
#
# Input JSON shape (produced by ai_xlsx_report in lib/ai.sh):
#   {
#     "vulnerabilities": [ { severity,name,cve_ref,category,affected_hosts,
#                            description,impact,remediation,cvss,tool_source } ],
#     "attack_chains":   [ { chain_name,initial_access,lateral_escalation,
#                            impact,findings_used,severity } ]
#   }
#
# Usage: xlsx_report.py --json findings.json -o report.xlsx
# Pure presentation: no formulas, so no LibreOffice recalc step is needed.
# ==========================================================================
import argparse
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- Palette (lifted from the template) ----------------------------------
NAVY = "1C2833"      # header background
ROWBLUE = "EBF5FB"   # zebra band
WHITE = "FFFFFF"
SUBTITLE = "2E4057"  # attack-paths subtitle band

SEV_FILL = {"CRITICAL": "C00000", "HIGH": "ED7D31", "MEDIUM": "FFC000",
            "LOW": "70AD47", "INFO": "5B9BD5"}
SEV_FONT = {"CRITICAL": "FFFFFF", "HIGH": "FFFFFF", "MEDIUM": "1C2833",
            "LOW": "FFFFFF", "INFO": "FFFFFF"}
SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}

FONT = "Arial"
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ATTACK_SUBTITLE = ("The following chains demonstrate realistic attacker paths "
                   "from initial access to full domain compromise based on "
                   "discovered vulnerabilities.")


def norm_sev(s):
    s = str(s or "").strip().upper()
    return s if s in SEV_FILL else "INFO"


def header_cell(cell, text, size=10):
    cell.value = text
    cell.font = Font(name=FONT, size=size, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def data_cell(cell, value, fill, *, bold=False, center=False, color=NAVY):
    cell.value = value
    cell.font = Font(name=FONT, size=9, bold=bold, color=color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center" if center else "top",
                               wrap_text=True)
    cell.border = BORDER


def sev_badge(cell, severity):
    sev = norm_sev(severity)
    cell.value = sev
    cell.font = Font(name=FONT, size=9, bold=True, color=SEV_FONT[sev])
    cell.fill = PatternFill("solid", fgColor=SEV_FILL[sev])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


# ---- Sheet 1: Vulnerabilities --------------------------------------------
def build_vulns(ws, vulns):
    headers = ["#", "Severity", "Vulnerability / Finding", "CVE / Ref", "Category",
               "Affected Host(s)", "Description", "Impact", "Remediation", "CVSS",
               "Tool / Source"]
    widths = [5, 8, 35, 12, 18, 28, 50, 45, 40, 12, 20]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        header_cell(ws.cell(1, i), h)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    vulns = sorted(vulns, key=lambda v: SEV_ORDER.get(norm_sev(v.get("severity")), 4))
    if not vulns:
        data_cell(ws.cell(2, 1), "—", WHITE, center=True)
        data_cell(ws.cell(2, 3), "No findings recorded for this engagement.", WHITE)
        return

    for idx, v in enumerate(vulns, start=1):
        r = idx + 1
        ws.row_dimensions[r].height = 80
        base = ROWBLUE if r % 2 == 0 else WHITE
        data_cell(ws.cell(r, 1), idx, base, bold=True, center=True)
        sev_badge(ws.cell(r, 2), v.get("severity"))
        data_cell(ws.cell(r, 3), v.get("name", ""), base, bold=True)
        data_cell(ws.cell(r, 4), v.get("cve_ref", ""), base, center=True)
        data_cell(ws.cell(r, 5), v.get("category", ""), base)
        data_cell(ws.cell(r, 6), v.get("affected_hosts", ""), base)
        data_cell(ws.cell(r, 7), v.get("description", ""), base)
        data_cell(ws.cell(r, 8), v.get("impact", ""), base)
        data_cell(ws.cell(r, 9), v.get("remediation", ""), base)
        data_cell(ws.cell(r, 10), v.get("cvss", ""), base, center=True)
        data_cell(ws.cell(r, 11), v.get("tool_source", ""), base)


# ---- Sheet 2: Attack Paths & Chains --------------------------------------
def build_chains(ws, chains):
    widths = {"A": 5, "B": 22, "C": 50, "D": 45, "E": 40, "F": 28, "G": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Banner (B1:G1) and subtitle (B2:G2) — fill every cell in the merge so the
    # band reads solid, then write text into the top-left.
    ws.merge_cells("B1:G1")
    ws.merge_cells("B2:G2")
    for col in range(1, 8):
        b = ws.cell(1, col); b.fill = PatternFill("solid", fgColor=NAVY); b.border = BORDER
        s = ws.cell(2, col); s.fill = PatternFill("solid", fgColor=SUBTITLE); s.border = BORDER
    ws.cell(1, 2).value = "ATTACK PATHS & EXPLOITATION CHAINS"
    ws.cell(1, 2).font = Font(name=FONT, size=14, bold=True, color=WHITE)
    ws.cell(1, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(2, 2).value = ATTACK_SUBTITLE
    ws.cell(2, 2).font = Font(name=FONT, size=9, color=WHITE)
    ws.cell(2, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    headers = ["#", "Chain Name", "Step 1 – Initial Access",
               "Step 2 – Lateral / Escalation", "Step 3 – Impact",
               "Findings Used", "Severity"]
    for i, h in enumerate(headers, 1):
        header_cell(ws.cell(3, i), h)
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    chains = sorted(chains, key=lambda c: SEV_ORDER.get(norm_sev(c.get("severity")), 4))
    if not chains:
        data_cell(ws.cell(4, 1), "—", WHITE, center=True)
        data_cell(ws.cell(4, 2), "No attack chains synthesised.", WHITE)
        return

    for idx, c in enumerate(chains, start=1):
        r = idx + 3
        ws.row_dimensions[r].height = 100
        base = ROWBLUE if r % 2 == 0 else WHITE
        data_cell(ws.cell(r, 1), idx, base, bold=True, center=True)
        data_cell(ws.cell(r, 2), c.get("chain_name", ""), base, bold=True)
        data_cell(ws.cell(r, 3), c.get("initial_access", ""), base)
        data_cell(ws.cell(r, 4), c.get("lateral_escalation", ""), base)
        data_cell(ws.cell(r, 5), c.get("impact", ""), base)
        data_cell(ws.cell(r, 6), c.get("findings_used", ""), base)
        sev_badge(ws.cell(r, 7), c.get("severity"))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", required=True, help="findings JSON from ai_xlsx_report")
    ap.add_argument("-o", "--out", required=True, help="output .xlsx path")
    args = ap.parse_args()

    try:
        with open(args.json, encoding="utf-8") as fh:
            data = json.load(fh)
    except Exception as e:
        print(f"xlsx_report: cannot read {args.json}: {e}", file=sys.stderr)
        return 1

    vulns = data.get("vulnerabilities", []) or []
    chains = data.get("attack_chains", []) or []

    wb = Workbook()
    build_vulns(wb.active, vulns)
    wb.active.title = "Vulnerabilities"
    build_chains(wb.create_sheet("Attack Paths & Chains"), chains)
    wb.save(args.out)
    print(f"xlsx_report: wrote {args.out} "
          f"({len(vulns)} findings, {len(chains)} chains)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
